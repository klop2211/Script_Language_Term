#!/usr/bin/python
# coding=utf-8

import sys
import time
import sqlite3
import telepot
from pprint import pprint
from urllib.request import urlopen
from bs4 import BeautifulSoup
import re
from datetime import date, datetime, timedelta
import traceback
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import requests
import xml.etree.ElementTree as ET


service_key = 'oWT9dq/7S6E11bbGOg2qY18HrHxLHQov6dy6WglV6AzFYqnOUC0YnQt/3GblAB1ygHHoWQpDuY0Wqc8grUD5oQ=='
TOKEN = '6246353591:AAGMZbi9hYtrDumgDzhVeahh51TOOplj9m4'
MAX_MSG_LENGTH = 300
baseurl = 'http://apis.data.go.kr/1360000/HealthWthrIdxServiceV3'
getPine = '/getPinePollenRiskIdxV3'
getOak = '/getOakPollenRiskIdxV3'
getWeed = '/getWeedPollenRiskIdxV3'
bot = telepot.Bot(TOKEN)

def getData(si, do, gun, what):
    # 지역 데이터
    workbook = load_workbook('지역별 지점코드(20230330).xlsx', data_only=True)
    sheet = workbook['최종 업데이트 파일_20230330']

    data = []

    ck = True
    for row in sheet.iter_rows(values_only=True):
        if row[2] != None and row[3] != None and row[4] != None:
            t = row[2] + ' ' + row[3] + ' ' + row[4]
            data.append(t)

        if ck:  # 맨 처음 데이터 버리기
            data.clear()
            ck = False

    workbook.close()

    now = datetime.now()
    before_12_hour = now - relativedelta(hours=12)
    string_now = now.strftime('%Y%m%d%H')
    string_18 = now.strftime('%Y%m%d18')
    string_06 = now.strftime('%Y%m%d06')

    if string_18 > string_now and string_now > string_06:
        now = string_now
    else:
        now = before_12_hour.strftime('%Y%m%d%H')

    queryParams = {'serviceKey': service_key, 'numOfRows': '3796', 'pageNo': '1', 'areaNo': '', 'time': now}

    dataList = []
    if what == '소나무':
        url = baseurl+getPine

        response = requests.get(url, params=queryParams)
        root = ET.fromstring(response.text)

        temp = []
        Pine_data = []

        # 읽어올  값
        for item in root.iter("item"):
            areaNo = item.findtext("areaNo")  # 지점코드
            date = item.findtext("date")  # 발표시간
            today = item.findtext("today")  # 오늘 예측값
            tomorrow = item.findtext("tomorrow")  # 내일 예측값
            dayaftertomorrow = item.findtext("dayaftertomorrow")  # 모레 예측값

            temp = [areaNo, date, today, tomorrow, dayaftertomorrow]
            Pine_data.append(temp)

        for i in range(len(data)):
            if data[i] == si + ' '  + do + ' ' + gun:
                dataList.append(Pine_data[i][2]) # 오늘
                dataList.append(Pine_data[i][3]) # 내일
                dataList.append(Pine_data[i][4]) # 모레

        return dataList

    elif what == '참나무':
        url = baseurl + getOak

        response = requests.get(url, params=queryParams)
        root = ET.fromstring(response.text)

        temp = []
        Oak_data = []

        # 읽어올  값
        for item in root.iter("item"):
            areaNo = item.findtext("areaNo")  # 지점코드
            date = item.findtext("date")  # 발표시간
            today = item.findtext("today")  # 오늘 예측값
            tomorrow = item.findtext("tomorrow")  # 내일 예측값
            dayaftertomorrow = item.findtext("dayaftertomorrow")  # 모레 예측값

            temp = [areaNo, date, today, tomorrow, dayaftertomorrow]
            Oak_data.append(temp)

        for i in range(len(data)):
            if data[i] == si + ' '  + do + ' ' + gun:
                dataList.append(Oak_data[i][2])  # 오늘
                dataList.append(Oak_data[i][3])  # 내일
                dataList.append(Oak_data[i][4])  # 모레

        return dataList

    elif what == '잡초류':
        url = baseurl + getWeed

        response = requests.get(url, params=queryParams)
        root = ET.fromstring(response.text)

        temp = []
        Weed_data = []

        # 읽어올  값
        for item in root.iter("item"):
            areaNo = item.findtext("areaNo")  # 지점코드
            date = item.findtext("date")  # 발표시간
            today = item.findtext("today")  # 오늘 예측값
            tomorrow = item.findtext("tomorrow")  # 내일 예측값
            dayaftertomorrow = item.findtext("dayaftertomorrow")  # 모레 예측값

            temp = [areaNo, date, today, tomorrow, dayaftertomorrow]
            Weed_data.append(temp)

        for i in range(len(data)):
            if data[i] == si + ' '  + do + ' ' + gun:
                dataList.append(Weed_data[i][2])  # 오늘
                dataList.append(Weed_data[i][3])  # 내일
                dataList.append(Weed_data[i][4])  # 모레

        return dataList
def sendMessage(user, msg):
    try:
        bot.sendMessage(user, msg)
    except:
        traceback.print_exc(file=sys.stdout)
