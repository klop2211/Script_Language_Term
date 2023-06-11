from datetime import datetime
from tkinter import *
import requests
import xml.etree.ElementTree as ET
import urllib.request
from openpyxl import load_workbook
import pickle
import os
from PIL import Image, ImageTk
import spam
from dateutil.relativedelta import relativedelta

# 지역 데이터
workbook = load_workbook('지역별 지점코드(20230330).xlsx', data_only=True)
sheet = workbook['최종 업데이트 파일_20230330']

data = []

ck = True
for row in sheet.iter_rows(values_only=True):
    if row[2] != None and row[3] != None and row[4] != None:
        t = row[2] + ' ' + row[3] + ' ' + row[4]
        data.append(t)

    if ck:  # 맨 처음 데이터 버리기
        data.clear()
        ck = False

workbook.close()

# 소나무, 참나무만 4~6 데이터를 제공
# 소나무
url = 'http://apis.data.go.kr/1360000/HealthWthrIdxServiceV3/getPinePollenRiskIdxV3'
service_key = "oWT9dq/7S6E11bbGOg2qY18HrHxLHQov6dy6WglV6AzFYqnOUC0YnQt/3GblAB1ygHHoWQpDuY0Wqc8grUD5oQ=="
now = datetime.now()
before_12_hour = now - relativedelta(hours=12)
string_now = now.strftime('%Y%m%d%H')
string_18 = now.strftime('%Y%m%d18')
string_06 = now.strftime('%Y%m%d06')

if string_18 > string_now and string_now > string_06:
    now = string_now
else:
    now = before_12_hour.strftime('%Y%m%d%H')


queryParams = {'serviceKey': service_key, 'numOfRows': '3796', 'pageNo': '1', 'areaNo': '', 'time': now}

response = requests.get(url, params=queryParams)
root = ET.fromstring(response.text)

temp = []
Pine_data = []

# 읽어올  값
for item in root.iter("item"):
    areaNo = item.findtext("areaNo") # 지점코드
    date = item.findtext("date") # 발표시간
    today = item.findtext("today") # 오늘 예측값
    tomorrow = item.findtext("tomorrow") # 내일 예측값
    dayaftertomorrow = item.findtext("dayaftertomorrow") # 모레 예측값

    temp = [areaNo, date, today, tomorrow, dayaftertomorrow]
    Pine_data.append(temp)

# 참나무
url = 'http://apis.data.go.kr/1360000/HealthWthrIdxServiceV3/getOakPollenRiskIdxV3'
queryParams = {'serviceKey': service_key, 'numOfRows': '3796', 'pageNo': '1', 'areaNo': '', 'time': now}

response = requests.get(url, params=queryParams)
root = ET.fromstring(response.text)

temp = []
Oak_data = []

# 읽어올  값
for item in root.iter("item"):
    areaNo = item.findtext("areaNo") # 지점코드
    date = item.findtext("date") # 발표시간
    today = item.findtext("today") # 오늘 예측값
    tomorrow = item.findtext("tomorrow") # 내일 예측값
    dayaftertomorrow = item.findtext("dayaftertomorrow") # 모레 예측값

    temp = [areaNo, date, today, tomorrow, dayaftertomorrow]
    Oak_data.append(temp)


# 잡초류

url = 'http://apis.data.go.kr/1360000/HealthWthrIdxServiceV3/getWeedsPollenRiskIdxV3'
queryParams = {'serviceKey': service_key, 'numOfRows': '3796', 'pageNo': '1', 'areaNo': '', 'time': now}

response = requests.get(url, params=queryParams)
root = ET.fromstring(response.text)

temp = []
Weed_data = []

# 읽어올  값
for item in root.iter("item"):
    areaNo = item.findtext("areaNo") # 지점코드
    date = item.findtext("date") # 발표시간
    today = item.findtext("today") # 오늘 예측값
    tomorrow = item.findtext("tomorrow") # 내일 예측값
    dayaftertomorrow = item.findtext("dayaftertomorrow") # 모레 예측값

    temp = [areaNo, date, today, tomorrow, dayaftertomorrow]
    Weed_data.append(temp)

# 구글 맵 API 키
MAP_API_KEY = "AIzaSyDldj_-4P3T4gKWdy6zqThReKArNFUXWAM"

class MainGUI():
    # 북마크 리스트 박스내의 아이템 더블클릭시 실행되는 함수
    def double_clickBookmark(self, event):
        # 선택된 지역에 대한 검색을 실행한다
        self.selected_item = self.bookmarkListbox.get(self.bookmarkListbox.curselection())
        print(f"You double-clicked: {self.selected_item}")

        wedo = -1
        kyoungdo = -1
        # 위도 경도 뽑아내야함
        for row in sheet.iter_rows(values_only=True):
            if row[2] != None and row[3] != None and row[4] != None:
                t = row[2] + ' ' + row[3] + ' ' + row[4]
                if self.selected_item == t:
                    wedo = row[14]
                    kyoungdo = row[13]

        if wedo != -1 and kyoungdo != -1:
            # 이 함수의 인자로 위도, 경도 넣어주면 됌
            self.drawMap(wedo, kyoungdo)


    # 검색 리스트 박스내의 아이템 더블클릭시 실행되는 함수
    def double_clickSearch(self, event):
        # 선택된 지역에 대한 검색을 실행한다
        self.selected_item = self.searchListbox.get(self.searchListbox.curselection())
        print(f"You double-clicked: {self.selected_item}")
        wedo = -1
        kyoungdo = -1
        # 위도 경도 뽑아내야함
        for row in sheet.iter_rows(values_only=True):
            if row[2] != None and row[3] != None and row[4] != None:
                t = row[2] + ' ' + row[3] + ' ' + row[4]
                if self.selected_item == t:
                    wedo = row[14]
                    kyoungdo = row[13]

        if wedo != -1 and kyoungdo != -1:
            # 이 함수의 인자로 위도, 경도 넣어주면 됌
            self.drawMap(wedo, kyoungdo)


    # 리스트 박스와 스크롤 바 길이 맞춤
    def configure_scrollbar(self, event):
        self.searchScrollbar.place_configure(height=event.height)
        self.bookmarkScrollbar.place_configure(height=event.height)

    # 검색 버튼이 눌렸을 때 작동하는 함수
    def commandSearch(self):
        # 지역검색이 이루어져야 하며 self.searchListbox.insert 함수로 리스트 박스를 채워야 한다
        word = self.searchEntry.get() # searchEntry에 있는 값 가져옴
        self.searchListbox.delete(0, END) # 리스트 박스 초기화
        for datas in data: # data 에서 하나 씩 가져와서
            if word in datas: # 검색한 단어가 하나라도 들어가 있으면
                self.searchListbox.insert(END, datas) # 리스트에 표현

    # 위도, 경도를 받아 지도를 그려주는 함수
    def drawMap(self, latitude=37.541, longitude=126.986):
        size = "250x250"
        zoom = 15
        url = f"https://maps.googleapis.com/maps/api/staticmap?center={latitude},{longitude}&zoom={zoom}&size={size}&key={MAP_API_KEY}"
        urllib.request.urlretrieve(url, "map.png")
        map_image = PhotoImage(file="map.png")
        self.mapLabel.configure(image=map_image)
        self.mapLabel.image = map_image

    # 종류, 리스트(오늘, 내일, 모레 3일치 값을 담은)
    def drawGraph(self, name, dataList):
        # 그래프 캔버스
        self.canvas = Canvas(self.window, width=250, height=250)
        self.canvas.place(x=300, y=550)
        width = 250 - 10
        height = 200 * 0.75
        maxCount = 0
        if dataList:
            maxCount = max(dataList)
            maxCount = int(maxCount)
        if maxCount == 0:
            maxCount = 1

        if dataList:
            for i  in range(3):
                dataList[i] = int(dataList[i])

        textlist = ['오늘', '내일', '모레']

        if not dataList:
            self.canvas.create_text(width / 2, 20, text = '자료 제공 기간이 아닙니다.')
            self.drawPhoto(name)
        else:
            for i in range(3):
                self.canvas.create_rectangle(i * width / 3 + 5, 200 - (height * dataList[i]/maxCount) - 5, (i + 1) * width / 3 + 5, 200 - 5)
                self.canvas.create_text(i * width / 3 + 5 + 0.5*width/3, 200 + 5, text=textlist[i])
            self.canvas.create_text(width / 2, 20, text=name)
            self.drawPhoto(name)

    # 꽃가루 종류
    def drawPhoto(self, name):
        # '문자열' + '.png' 를 리턴하는 C++ 연동
        filename = spam.plusPNG(name)
        image = PhotoImage(file=f'image/{filename}')
        self.photoLabel.configure(image=image)
        self.photoLabel.image = image

    # 현재 searchListbox 에서 선택된 값을 즐겨찾기에 추가한다
    def commandBookmarkInsert(self):
        selected_item = self.searchListbox.get(self.searchListbox.curselection())
        if type(selected_item) == str:
            self.bookmarkListbox.insert(END, selected_item)
            with open('bookmark.pickle', 'wb') as f:
                pickle.dump(self.bookmarkListbox.get(0, END), f)

    # 즐겨찾기 삭제
    def commandBookmarkDelete(self):
        self.bookmarkListbox.delete(self.bookmarkListbox.curselection())
        with open('bookmark.pickle', 'wb') as f:
            pickle.dump(self.bookmarkListbox.get(0, END), f)

    # 소나무에 대한 정보 표시
    def commandSonamu(self):
        dataList = []

        # 여기에 값을 넣어주세요
        for i in range(len(data)):
            if data[i] == self.selected_item:
                dataList.append(Pine_data[i][2]) # 오늘
                dataList.append(Pine_data[i][3]) # 내일
                dataList.append(Pine_data[i][4]) # 모레
        # 현재 선택된 지역은 self.selected_item 입니다
        self.drawGraph('소나무', dataList)

    # 참나무에 대한 정보 표시
    def commandChamnamu(self):
        # 여기에 값을 넣어주세요
        dataList = []

        if Oak_data:
            for i in range(len(data)):
                if data[i] == self.selected_item:
                    dataList.append(Oak_data[i][2]) # 오늘
                    dataList.append(Oak_data[i][3]) # 내일
                    dataList.append(Oak_data[i][4]) # 모레
        # 현재 선택된 지역은 self.selected_item 입니다
        self.drawGraph('참나무', dataList)

    # 잡초류에 대한 정보 표시
    def commandJapchoryu(self):
        # 여기에 값을 넣어주세요
        dataList = []

        if Weed_data:
            for i in range(len(data)):
                if data[i] == self.selected_item:
                    dataList.append(Weed_data[i][2]) # 오늘
                    dataList.append(Weed_data[i][3]) # 내일
                    dataList.append(Weed_data[i][4]) # 모레

        # 현재 선택된 지역은 self.selected_item 입니다
        self.drawGraph('잡초류', dataList)

    # UI 설정
    def setUI(self):

        # 검색 버튼과 검색창
        self.searchEntry = Entry(self.window, font=('arial', 30))
        self.searchButton = Button(self.window, text='검색', font=('arial', 20), command=self.commandSearch)
        self.searchEntry.place(x=0)
        self.searchButton.place(x=500)

        # 검색 결과창
        self.searchScrollbar = Scrollbar(self.window)
        self.searchScrollbar.place(x=220, y=70)

        self.searchListbox = Listbox(self.window, font=('arial', 10), width=25, height=7, yscrollcommand=self.searchScrollbar.set)
        self.searchListbox.place(x=20, y=70)
        self.searchListbox.bind('<Configure>', self.configure_scrollbar)
        self.searchListbox.bind("<Double-Button-1>", self.double_clickSearch)
        self.searchScrollbar.config(command=self.searchListbox.yview)

        # 북마크 목록
        self.bookmarkScrollbar = Scrollbar(self.window)
        self.bookmarkScrollbar.place(x=470, y=70)

        self.bookmarkListbox = Listbox(self.window, font=('arial', 10), width=25, height=7, yscrollcommand=self.bookmarkScrollbar.set)
        self.bookmarkListbox.place(x=270, y=70)
        self.bookmarkListbox.bind('<Configure>', self.configure_scrollbar)
        self.bookmarkListbox.bind("<Double-Button-1>", self.double_clickBookmark)

        # 저장된 즐겨찾기 불러옴
        if os.path.exists('bookmark.pickle'):
            with open('bookmark.pickle', 'rb') as f:
                readBookmark = pickle.load(f)
            self.bookmarkListbox.insert(END, *readBookmark)
        self.bookmarkScrollbar.config(command=self.bookmarkListbox.yview)

        self.bookmarkButton = Button(self.window, text='즐겨찾기 추가', command=self.commandBookmarkInsert)
        self.bookmarkButton.place(x=500, y=80)
        self.bookmarkButton = Button(self.window, text='즐겨찾기 삭제', command=self.commandBookmarkDelete)
        self.bookmarkButton.place(x=500, y=130)

        # 지도 그리기용 라벨
        self.mapLabel = Label(self.window)
        self.mapLabel.place(x=20, y=250)

        # 정보 검색용 버튼
        self.sonamuButton = Button(self.window, text='소나무', command=self.commandSonamu)
        self.sonamuButton.place(x=300, y=250)
        self.chamnamuButton = Button(self.window, text='참나무', command=self.commandChamnamu)
        self.chamnamuButton.place(x=400, y=250)
        self.japchoryuButton = Button(self.window, text='잡초류', command=self.commandJapchoryu)
        self.japchoryuButton.place(x=500, y=250)

        # 꽃가루 종류 사진 라벨
        self.photoLabel = Label(self.window, width=250, height=250)
        self.photoLabel.place(x=300, y=300)


    def __init__(self):
        self.window = Tk()
        self.window.title("플라워타임")
        self.window.geometry("600x800")
        self.setUI()
        self.window.mainloop()

# aaaa
MainGUI()